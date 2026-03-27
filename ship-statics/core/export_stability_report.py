# ============================================================
# core/export_stability_report.py — 稳性校核报告 Excel 导出模块
#
# 阶段6 第三优先级
#
# 功能：
#   1. 生成完整的稳性校核报告 Excel
#   2. 包含工况参数、计算结果、衡准指标、判定结论
#   3. 格式规范，可直接用于课程设计
#
# ============================================================

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from typing import Dict, List, Optional
import os

# ══════════════════════════════════════════════════════════
# Excel 导出
# ══════════════════════════════════════════════════════════

class StabilityReportExporter:
    """稳性校核报告 Excel 导出类"""
    
    def __init__(self):
        """初始化导出器"""
        self.wb = None
        self.ws = None
    
    def create_single_condition_report(self, condition_name: str, 
                                      analysis_result: Dict,
                                      judgment_result: Dict,
                                      output_path: str) -> str:
        """
        创建单工况稳性校核报告 Excel
        
        参数：
            condition_name: 工况名称
            analysis_result: 分析结果
            judgment_result: 判定结果
            output_path: 输出路径
        
        返回：
            输出文件路径
        """
        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active
        self.ws.title = "稳性校核报告"
        
        # 设置列宽
        self.ws.column_dimensions['A'].width = 20
        self.ws.column_dimensions['B'].width = 15
        self.ws.column_dimensions['C'].width = 15
        self.ws.column_dimensions['D'].width = 20
        
        row = 1
        
        # ── 标题 ──
        self.ws.merge_cells(f'A{row}:D{row}')
        title_cell = self.ws[f'A{row}']
        title_cell.value = '船舶稳性校核报告'
        title_cell.font = Font(name='宋体', size=16, bold=True, color='FFFFFF')
        title_cell.fill = PatternFill(start_color='1a4a8a', end_color='1a4a8a', fill_type='solid')
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        self.ws.row_dimensions[row].height = 25
        row += 1
        
        # ── 工况信息 ──
        row = self._add_section_header(row, '工况信息')
        
        loading = analysis_result.get('loading', {})
        self._add_row(row, '工况名称', condition_name)
        row += 1
        self._add_row(row, '总重量 Δ (t)', f"{loading.get('total_weight', 0):.2f}")
        row += 1
        self._add_row(row, '重心纵向 Xg (m)', f"{loading.get('xg', 0):.2f}")
        row += 1
        self._add_row(row, '重心横向 Yg (m)', f"{loading.get('yg', 0):.2f}")
        row += 1
        self._add_row(row, '重心垂向 Zg (m)', f"{loading.get('zg', 0):.2f}")
        row += 1
        
        # ── 浮态计算结果 ──
        row = self._add_section_header(row, '浮态计算结果')
        
        floating_state = analysis_result.get('floating_state', {})
        self._add_row(row, '平均吃水 (m)', f"{floating_state.get('draft_mean', 0):.2f}")
        row += 1
        self._add_row(row, '首吃水 (m)', f"{floating_state.get('draft_fwd', 0):.2f}")
        row += 1
        self._add_row(row, '尾吃水 (m)', f"{floating_state.get('draft_aft', 0):.2f}")
        row += 1
        self._add_row(row, '纵倾 (m)', f"{floating_state.get('trim', 0):.2f}")
        row += 1
        
        # ── 稳性指标 ──
        row = self._add_section_header(row, '稳性指标')
        
        gz_curve = analysis_result.get('gz_curve', {})
        indicators = judgment_result.get('indicators', {})
        
        self._add_row(row, '初稳性高度 GM (m)', f"{gz_curve.get('GM', 0):.4f}")
        row += 1
        self._add_row(row, '最大复原力臂 GZ_max (m)', f"{indicators.get('GZ_max', 0):.4f}")
        row += 1
        self._add_row(row, '最大复原力臂角度 θ_max_gz (°)', f"{indicators.get('theta_max_gz', 0):.1f}")
        row += 1
        self._add_row(row, '稳性消失角 θ_vanish (°)', f"{indicators.get('theta_vanish', 0):.1f}")
        row += 1
        self._add_row(row, '稳性衡准数 K', f"{indicators.get('K', 0):.4f}")
        row += 1
        
        # ── 规范衡准校核 ──
        row = self._add_section_header(row, '规范衡准校核（《国内航行海船法定检验技术规则》）')
        
        judgments = judgment_result.get('judgments', {})
        
        # 表头
        self.ws[f'A{row}'] = '衡准项'
        self.ws[f'B{row}'] = '计算值'
        self.ws[f'C{row}'] = '规范要求'
        self.ws[f'D{row}'] = '判定结果'
        
        for col in ['A', 'B', 'C', 'D']:
            cell = self.ws[f'{col}{row}']
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='2e7dd1', end_color='2e7dd1', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        row += 1
        
        # 衡准项数据
        criteria_names = {
            'GM': '初稳性高度',
            'GZ_max': '最大复原力臂',
            'theta_max_gz': '最大复原力臂角度',
            'theta_vanish': '稳性消失角',
            'K': '稳性衡准数'
        }
        
        for key, name in criteria_names.items():
            if key in judgments:
                judgment = judgments[key]
                passed = judgment.get('passed', False)
                
                self.ws[f'A{row}'] = name
                self.ws[f'B{row}'] = f"{judgment.get('value', 0):.4f}"
                self.ws[f'C{row}'] = f"≥ {judgment.get('limit', 0)}"
                self.ws[f'D{row}'] = '✓ 通过' if passed else '✗ 不通过'
                
                # 不通过时标红
                if not passed:
                    for col in ['A', 'B', 'C', 'D']:
                        self.ws[f'{col}{row}'].fill = PatternFill(start_color='ffcccc', end_color='ffcccc', fill_type='solid')
                
                row += 1
        
        # ── 总体结论 ──
        row += 1
        row = self._add_section_header(row, '总体结论')
        
        overall_pass = judgment_result.get('overall_pass', False)
        conclusion = '✓ 稳性合格' if overall_pass else '✗ 稳性不合格'
        
        self.ws.merge_cells(f'A{row}:D{row}')
        conclusion_cell = self.ws[f'A{row}']
        conclusion_cell.value = conclusion
        conclusion_cell.font = Font(size=12, bold=True, color='FFFFFF')
        conclusion_cell.fill = PatternFill(
            start_color='27ae60' if overall_pass else 'e74c3c',
            end_color='27ae60' if overall_pass else 'e74c3c',
            fill_type='solid'
        )
        conclusion_cell.alignment = Alignment(horizontal='center', vertical='center')
        self.ws.row_dimensions[row].height = 20
        row += 1
        
        # 不满足项
        if not overall_pass:
            failed_items = judgment_result.get('failed_items', [])
            if failed_items:
                self.ws.merge_cells(f'A{row}:D{row}')
                failed_cell = self.ws[f'A{row}']
                failed_cell.value = f'不满足项：{", ".join(failed_items)}'
                failed_cell.font = Font(size=11, color='FFFFFF')
                failed_cell.fill = PatternFill(start_color='e74c3c', end_color='e74c3c', fill_type='solid')
                failed_cell.alignment = Alignment(horizontal='center', vertical='center')
                row += 1
        
        # ── 页脚 ──
        row += 2
        self.ws.merge_cells(f'A{row}:D{row}')
        footer_cell = self.ws[f'A{row}']
        footer_cell.value = f'生成时间：{datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
        footer_cell.font = Font(size=9, italic=True, color='666666')
        footer_cell.alignment = Alignment(horizontal='right')
        
        # 保存文件
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        self.wb.save(output_path)
        
        return output_path
    
    def create_all_conditions_report(self, all_results: Dict, 
                                    output_path: str) -> str:
        """
        创建全工况稳性校核报告 Excel
        
        参数：
            all_results: {工况名: {analysis, judgment}}
            output_path: 输出路径
        
        返回：
            输出文件路径
        """
        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active
        self.ws.title = "全工况校核"
        
        # 设置列宽
        for col in range(1, 10):
            self.ws.column_dimensions[get_column_letter(col)].width = 15
        
        row = 1
        
        # ── 标题 ──
        self.ws.merge_cells(f'A{row}:I{row}')
        title_cell = self.ws[f'A{row}']
        title_cell.value = '全工况稳性校核报告'
        title_cell.font = Font(name='宋体', size=16, bold=True, color='FFFFFF')
        title_cell.fill = PatternFill(start_color='1a4a8a', end_color='1a4a8a', fill_type='solid')
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        self.ws.row_dimensions[row].height = 25
        row += 2
        
        # ── 表头 ──
        headers = ['工况名称', '判定结果', 'GM (m)', 'GZ_max (m)', 'θ_max_gz (°)', 
                  'θ_vanish (°)', 'K 值', '不满足项', '规范条款']
        
        for col, header in enumerate(headers, 1):
            cell = self.ws.cell(row=row, column=col)
            cell.value = header
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='2e7dd1', end_color='2e7dd1', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        row += 1
        
        # ── 数据行 ──
        for condition_name, result in all_results.items():
            judgment = result.get('judgment', {})
            indicators = judgment.get('indicators', {})
            passed = judgment.get('overall_pass', False)
            failed_items = judgment.get('failed_items', [])
            
            self.ws.cell(row=row, column=1).value = condition_name
            self.ws.cell(row=row, column=2).value = '✓ 合格' if passed else '✗ 不合格'
            self.ws.cell(row=row, column=3).value = f"{indicators.get('GM', 0):.4f}"
            self.ws.cell(row=row, column=4).value = f"{indicators.get('GZ_max', 0):.4f}"
            self.ws.cell(row=row, column=5).value = f"{indicators.get('theta_max_gz', 0):.1f}"
            self.ws.cell(row=row, column=6).value = f"{indicators.get('theta_vanish', 0):.1f}"
            self.ws.cell(row=row, column=7).value = f"{indicators.get('K', 0):.4f}"
            self.ws.cell(row=row, column=8).value = ', '.join(failed_items) if failed_items else '无'
            self.ws.cell(row=row, column=9).value = '规则第4篇第3章'
            
            # 不通过时标红
            if not passed:
                for col in range(1, 10):
                    self.ws.cell(row=row, column=col).fill = PatternFill(
                        start_color='ffcccc', end_color='ffcccc', fill_type='solid'
                    )
            
            row += 1
        
        # ── 总体统计 ──
        row += 1
        total = len(all_results)
        passed_count = sum(1 for r in all_results.values() if r.get('judgment', {}).get('overall_pass', False))
        failed_count = total - passed_count
        
        self.ws.merge_cells(f'A{row}:I{row}')
        summary_cell = self.ws[f'A{row}']
        summary_cell.value = f'总计：{total} 工况 | 合格：{passed_count} | 不合格：{failed_count}'
        summary_cell.font = Font(size=11, bold=True, color='FFFFFF')
        summary_cell.fill = PatternFill(start_color='1a4a8a', end_color='1a4a8a', fill_type='solid')
        summary_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 保存文件
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        self.wb.save(output_path)
        
        return output_path
    
    def _add_section_header(self, row: int, title: str) -> int:
        """添加章节标题"""
        self.ws.merge_cells(f'A{row}:D{row}')
        header_cell = self.ws[f'A{row}']
        header_cell.value = title
        header_cell.font = Font(bold=True, color='FFFFFF', size=11)
        header_cell.fill = PatternFill(start_color='2e7dd1', end_color='2e7dd1', fill_type='solid')
        header_cell.alignment = Alignment(horizontal='left', vertical='center')
        self.ws.row_dimensions[row].height = 18
        return row + 1
    
    def _add_row(self, row: int, label: str, value: str):
        """添加数据行"""
        self.ws[f'A{row}'] = label
        self.ws[f'B{row}'] = value
        
        # 样式
        for col in ['A', 'B']:
            cell = self.ws[f'{col}{row}']
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            cell.alignment = Alignment(horizontal='left', vertical='center')
        
        self.ws[f'A{row}'].font = Font(bold=True)


# ══════════════════════════════════════════════════════════
# 测试函数
# ══════════════════════════════════════════════════════════

if __name__ == '__main__':
    import sys
    sys.stdout.reconfigure(encoding='utf-8')
    
    print('=' * 70)
    print('稳性校核报告 Excel 导出模块测试')
    print('=' * 70)
    
    # 模拟数据
    mock_analysis = {
        'condition_name': '满载出港',
        'loading': {
            'total_weight': 5200,
            'xg': 42.5,
            'yg': 0,
            'zg': 5.1
        },
        'floating_state': {
            'draft_mean': 5.80,
            'draft_fwd': 5.80,
            'draft_aft': 5.80,
            'trim': 0.0
        },
        'gz_curve': {
            'GM': 0.8234
        }
    }
    
    mock_judgment = {
        'overall_pass': True,
        'indicators': {
            'GM': 0.8234,
            'GZ_max': 0.42,
            'theta_max_gz': 30.0,
            'theta_vanish': 85.0,
            'K': 1.2
        },
        'judgments': {
            'GM': {'passed': True, 'value': 0.8234, 'limit': 0.15},
            'GZ_max': {'passed': True, 'value': 0.42, 'limit': 0.20},
            'theta_max_gz': {'passed': True, 'value': 30.0, 'limit': 25.0},
            'theta_vanish': {'passed': True, 'value': 85.0, 'limit': 55.0},
            'K': {'passed': True, 'value': 1.2, 'limit': 1.0}
        },
        'failed_items': []
    }
    
    # 创建导出器
    exporter = StabilityReportExporter()
    
    # 导出单工况报告
    print('\n【导出单工况报告】')
    output_path = 'F:/ship-statics/outputs/稳性校核报告_满载出港.xlsx'
    exporter.create_single_condition_report('满载出港', mock_analysis, mock_judgment, output_path)
    print(f'✓ 单工况报告已导出：{output_path}')
    
    print('\n✓ 稳性校核报告 Excel 导出模块测试通过!')
