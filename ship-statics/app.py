# ============================================================
# app.py — Flask 主入口
#
# 船舶静力学课程设计自动计算系统（阶段4：曲线绘制）
# 武汉理工大学 船舶与海洋工程学院
#
# 阶段1 功能（完全保留）：手动输入 → 梯形法计算 → Excel 导出
# 阶段2 功能（完全保留）：Excel 型值表上传 + 数据校验
# 阶段3 功能（完全保留）：多吃水全参数计算（14项）
# 阶段4 新增功能：
#   - 邦戎曲线计算（/calc_bonjean）
#   - 静水力曲线绘制（/plot_hydrostatics）
#   - 邦戎曲线绘制（/plot_bonjean）
#   - 曲线预览（/preview_hydrostatics、/preview_bonjean）
#   - 曲线文件下载（/download_plot/<filename>）
#
# 启动方式：python app.py
# ============================================================
#   或双击 start.bat
# ============================================================

from flask import Flask, render_template, request, jsonify, send_file
from flask_cors import CORS
import os
import traceback
from datetime import datetime

# ── 阶段1 模块（不修改）──────────────────────────────────
from core.calculator import calculate_hydrostatics
from core.exporter   import export_to_excel

# ── 阶段2 新增模块 ────────────────────────────────────────
from core.excel_parser       import parse_offsets_excel
from core.template_generator import generate_template

# ── 阶段3 新增模块 ────────────────────────────────────────
from core.hydrostatics_full import calc_hydrostatics_table
from core.exporter_full     import export_hydrostatics_excel

# ── 阶段4 新增模块 ────────────────────────────────────────
from core.bonjean  import calc_bonjean_table
from core.plotter  import (plot_hydrostatics, plot_bonjean,
                            plot_hydrostatics_preview, plot_bonjean_preview)

# ── 阶段5 新增模块 ────────────────────────────────────────
from core.stability  import calc_gz_curve, calc_gz_family
from core.plotter_gz import plot_gz_curve, plot_gz_family, plot_gz_preview

# ── 阶段6 新增模块 ────────────────────────────────────────
from core.loading_condition import LoadingConditionManager
from core.floating_stability import LoadingConditionStability, StabilityIndicators
from core.stability_criteria import StabilityJudgment
from core.loading_stability_analysis import LoadingStabilityAnalysis
from core.export_stability_report import StabilityReportExporter

# ── 阶段7 新增模块 ────────────────────────────────────────
from core.word_report_generator import CourseDesignReportGenerator

# ── 阶段7 新增模块 ────────────────────────────────────────
from core.word_report_generator import CourseDesignReportGenerator

# ── 创建 Flask 应用 ──────────────────────────────────────
app = Flask(__name__)
CORS(app)  # 启用跨域支持，解决 Failed to fetch 问题
app.config["MAX_CONTENT_LENGTH"] = 50 * 1024 * 1024  # 最大上传 50MB

# 目录
OUTPUT_DIR = os.path.join(os.path.dirname(__file__), "outputs")
UPLOAD_DIR = os.path.join(os.path.dirname(__file__), "uploads")
os.makedirs(OUTPUT_DIR, exist_ok=True)
os.makedirs(UPLOAD_DIR, exist_ok=True)

# 预先生成模板文件（启动时生成一次，避免每次请求都生成）
TEMPLATE_PATH = os.path.join(OUTPUT_DIR, "型值表模板.xlsx")


# ══════════════════════════════════════════════════════════
# 阶段1 路由（完全保留，代码未修改）
# ══════════════════════════════════════════════════════════

@app.route("/")
def index():
    """返回前端单页应用"""
    return render_template("index.html")


@app.route("/calculate", methods=["POST"])
def calculate():
    """
    接收前端提交的型值表数据，执行静水力计算，返回结果。
    （阶段1原有路由，未修改）

    请求体（JSON）：
        {
            "draft": 5.0,
            "stations": [-5,-4,...,5],
            "half_breadths": [0,1.2,...,0]
        }
    """
    try:
        data = request.get_json()
        if not data:
            return jsonify({"success": False, "msg": "未收到数据"})

        draft         = float(data.get("draft", 0))
        stations      = [float(x) for x in data.get("stations", [])]
        half_breadths = [float(y) for y in data.get("half_breadths", [])]

        result = calculate_hydrostatics(stations, half_breadths, draft)
        result["stations"]      = stations
        result["half_breadths"] = half_breadths

        return jsonify({"success": True, "result": result})

    except ValueError as e:
        return jsonify({"success": False, "msg": f"输入数据错误：{str(e)}"})
    except Exception as e:
        return jsonify({
            "success": False,
            "msg": f"计算出错：{str(e)}",
            "trace": traceback.format_exc()
        })


@app.route("/export_excel", methods=["POST"])
def export_excel():
    """
    接收计算结果，生成 Excel 文件并返回下载。
    （阶段1原有路由，未修改）
    """
    try:
        data          = request.get_json()
        result        = data.get("result", {})
        stations      = data.get("stations", [])
        half_breadths = data.get("half_breadths", [])

        timestamp   = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = os.path.join(OUTPUT_DIR, f"静水力计算_{timestamp}.xlsx")

        export_to_excel(result, stations, half_breadths, output_path)

        return send_file(
            output_path,
            as_attachment=True,
            download_name=f"静水力计算_{timestamp}.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        return jsonify({
            "success": False,
            "msg": f"导出失败：{str(e)}",
            "trace": traceback.format_exc()
        })


# ══════════════════════════════════════════════════════════
# 阶段2 新增路由
# ══════════════════════════════════════════════════════════

@app.route("/upload_offsets", methods=["POST"])
def upload_offsets():
    """
    【阶段2新增】接收上传的型值表 Excel 文件，解析并校验。

    请求：multipart/form-data，字段名 "file"

    返回（JSON）：
        {
            "success": true,
            "data": {
                "stations":    [-5, -4, ..., 5],
                "waterlines":  [0.0, 0.5, ..., 8.0],
                "offsets":     [[...], ...],   // offsets[wl_idx][sta_idx]
                "n_stations":  11,
                "n_waterlines": 17,
                "L":           10.0,
                "station_spacing": 1.0,
                "validation": {
                    "warnings":    [...],
                    "errors":      [...],
                    "corrections": [...],
                    "is_valid":    true
                }
            }
        }
    """
    if "file" not in request.files:
        return jsonify({"success": False, "msg": "未收到文件，请选择 .xlsx 文件上传"})

    f = request.files["file"]
    if not f.filename:
        return jsonify({"success": False, "msg": "文件名为空"})

    # 检查文件扩展名
    if not f.filename.lower().endswith(".xlsx"):
        return jsonify({"success": False, "msg": "仅支持 .xlsx 格式，请另存为 xlsx 后上传"})

    # 保存到 uploads 目录
    save_path = os.path.join(UPLOAD_DIR, "offsets_upload.xlsx")
    f.save(save_path)

    try:
        # 调用阶段2解析模块
        data = parse_offsets_excel(save_path)

        # 将 numpy float 转为 Python float（JSON 序列化需要）
        data["stations"]   = [float(x) for x in data["stations"]]
        data["waterlines"] = [float(z) for z in data["waterlines"]]
        data["offsets"]    = [[float(y) for y in row] for row in data["offsets"]]

        return jsonify({"success": True, "data": data})

    except ValueError as e:
        return jsonify({"success": False, "msg": f"解析失败：{str(e)}"})
    except Exception as e:
        return jsonify({
            "success": False,
            "msg": f"解析出错：{str(e)}",
            "trace": traceback.format_exc()
        })


@app.route("/download_template", methods=["GET"])
def download_template():
    """
    【阶段2新增】下载型值表 Excel 模板文件。

    返回：Excel 文件下载
    """
    try:
        # 每次请求都重新生成，确保模板是最新的
        generate_template(TEMPLATE_PATH, example=True)

        return send_file(
            TEMPLATE_PATH,
            as_attachment=True,
            download_name="型值表模板（含示例数据）.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception as e:
        return jsonify({"success": False, "msg": f"模板生成失败：{str(e)}"})


@app.route("/download_template_blank", methods=["GET"])
def download_template_blank():
    """
    【阶段2新增】下载空白型值表模板（不含示例数据）。
    """
    try:
        blank_path = os.path.join(OUTPUT_DIR, "型值表模板（空白）.xlsx")
        generate_template(blank_path, example=False)

        return send_file(
            blank_path,
            as_attachment=True,
            download_name="型值表模板（空白）.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception as e:
        return jsonify({"success": False, "msg": f"模板生成失败：{str(e)}"})


# ══════════════════════════════════════════════════════════
# 阶段3 新增路由
# ══════════════════════════════════════════════════════════

@app.route("/calc_table", methods=["POST"])
def calc_table():
    """
    【阶段3新增】多吃水遍历计算，返回完整静水力计算表。

    请求体（JSON）：
        {
            "offsets_data": {              // 型值表数据（来自上传或手动输入）
                "stations":    [...],
                "waterlines":  [...],
                "offsets":     [[...], ...]
            },
            "d_min":  0.5,                 // 最小吃水（m），可选
            "d_max":  8.0,                 // 最大吃水（m），可选
            "d_step": 0.5,                 // 吃水步长（m），默认 0.5
            "method": "trapz"              // 积分方法：trapz / simpson
        }

    返回（JSON）：
        {
            "success": true,
            "table": {
                "rows":     [...],   // 每行为一个吃水的全部参数
                "columns":  [...],   // 列定义
                "n_drafts": 17,
                ...
            }
        }
    """
    try:
        data = request.get_json()
        if not data:
            return jsonify({"success": False, "msg": "未收到数据"})

        offsets_data = data.get("offsets_data")
        if not offsets_data:
            return jsonify({"success": False, "msg": "缺少型值表数据（offsets_data）"})

        d_min  = data.get("d_min",  None)
        d_max  = data.get("d_max",  None)
        d_step = float(data.get("d_step", 0.5))
        method = data.get("method", "trapz")

        if d_min  is not None: d_min  = float(d_min)
        if d_max  is not None: d_max  = float(d_max)

        # 调用阶段3核心计算
        table = calc_hydrostatics_table(
            offsets_data=offsets_data,
            d_min=d_min,
            d_max=d_max,
            d_step=d_step,
            method=method
        )

        return jsonify({"success": True, "table": table})

    except ValueError as e:
        return jsonify({"success": False, "msg": f"参数错误：{str(e)}"})
    except Exception as e:
        return jsonify({
            "success": False,
            "msg": f"计算出错：{str(e)}",
            "trace": traceback.format_exc()
        })


@app.route("/export_hydrostatics", methods=["POST"])
def export_hydrostatics():
    """
    【阶段3新增】导出完整多吃水静水力计算表为 Excel。

    请求体（JSON）：
        {
            "table": { ... }   // calc_table() 返回的 table 对象
        }
    """
    try:
        data  = request.get_json()
        table = data.get("table")
        if not table:
            return jsonify({"success": False, "msg": "缺少计算表数据"})

        timestamp   = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = os.path.join(OUTPUT_DIR, f"静水力计算表_{timestamp}.xlsx")

        export_hydrostatics_excel(table, output_path)

        return send_file(
            output_path,
            as_attachment=True,
            download_name=f"静水力计算表_{timestamp}.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        return jsonify({
            "success": False,
            "msg": f"导出失败：{str(e)}",
            "trace": traceback.format_exc()
        })


# ══════════════════════════════════════════════════════════
# 阶段4 新增路由：曲线计算与绘制
# ══════════════════════════════════════════════════════════

# 曲线输出目录
PLOT_DIR = os.path.join(OUTPUT_DIR, "plots")
os.makedirs(PLOT_DIR, exist_ok=True)


@app.route("/calc_bonjean", methods=["POST"])
def calc_bonjean():
    """
    【阶段4新增】计算邦戎曲线数据表。

    请求体（JSON）：
        {
            "offsets_data": { stations, waterlines, offsets },
            "method": "trapz"
        }

    返回（JSON）：
        {
            "success": true,
            "bonjean": {
                "stations":    [...],
                "waterlines":  [...],
                "A":           [[...], ...],   // 横剖面面积 [sta][wl]
                "S":           [[...], ...],   // 面积静矩   [sta][wl]
                "n_stations":  11,
                "n_waterlines": 17
            }
        }
    """
    try:
        data = request.get_json()
        offsets_data = data.get("offsets_data")
        if not offsets_data:
            return jsonify({"success": False, "msg": "缺少型值表数据"})

        method = data.get("method", "trapz")
        bonjean = calc_bonjean_table(offsets_data, method=method)
        return jsonify({"success": True, "bonjean": bonjean})

    except Exception as e:
        return jsonify({
            "success": False,
            "msg": f"邦戎曲线计算出错：{str(e)}",
            "trace": traceback.format_exc()
        })


@app.route("/preview_hydrostatics", methods=["POST"])
def preview_hydrostatics():
    """
    【阶段4新增】生成静水力曲线预览图（base64 PNG，低分辨率）。

    请求体（JSON）：{ "table": { ... } }
    返回（JSON）：{ "success": true, "image": "base64..." }
    """
    try:
        data  = request.get_json()
        table = data.get("table")
        if not table:
            return jsonify({"success": False, "msg": "缺少计算表数据"})

        b64 = plot_hydrostatics_preview(table)
        return jsonify({"success": True, "image": b64})

    except Exception as e:
        return jsonify({
            "success": False,
            "msg": f"预览生成失败：{str(e)}",
            "trace": traceback.format_exc()
        })


@app.route("/preview_bonjean", methods=["POST"])
def preview_bonjean():
    """
    【阶段4新增】生成邦戎曲线预览图（base64 PNG，低分辨率）。

    请求体（JSON）：{ "bonjean": { ... } }
    返回（JSON）：{ "success": true, "image": "base64..." }
    """
    try:
        data    = request.get_json()
        bonjean = data.get("bonjean")
        if not bonjean:
            return jsonify({"success": False, "msg": "缺少邦戎曲线数据"})

        b64 = plot_bonjean_preview(bonjean)
        return jsonify({"success": True, "image": b64})

    except Exception as e:
        return jsonify({
            "success": False,
            "msg": f"预览生成失败：{str(e)}",
            "trace": traceback.format_exc()
        })


@app.route("/export_plots", methods=["POST"])
def export_plots():
    """
    【阶段4新增】生成高清曲线图（PNG 300DPI + SVG），打包为 ZIP 下载。

    请求体（JSON）：
        {
            "table":     { ... },   // 静水力计算表（必须）
            "bonjean":   { ... },   // 邦戎曲线数据（可选）
            "ship_name": "示例船",  // 船名（可选）
            "formats":   ["png", "svg"]  // 导出格式（可选，默认两种都导出）
        }
    """
    try:
        import zipfile
        data      = request.get_json()
        table     = data.get("table")
        bonjean   = data.get("bonjean")
        ship_name = data.get("ship_name", "")
        formats   = data.get("formats", ["png", "svg"])

        if not table:
            return jsonify({"success": False, "msg": "缺少静水力计算表数据"})

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        plot_dir  = os.path.join(PLOT_DIR, timestamp)
        os.makedirs(plot_dir, exist_ok=True)

        # ── 绘制静水力曲线 ────────────────────────────────
        hydro_paths = plot_hydrostatics(table, plot_dir, ship_name=ship_name)

        # ── 绘制邦戎曲线（如果有数据）────────────────────
        bonjean_paths = {}
        if bonjean:
            bonjean_paths = plot_bonjean(bonjean, plot_dir, ship_name=ship_name)

        # ── 打包为 ZIP ────────────────────────────────────
        zip_path = os.path.join(OUTPUT_DIR, f"曲线图_{timestamp}.zip")
        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zf:
            for fmt in formats:
                if fmt in hydro_paths:
                    zf.write(hydro_paths[fmt],
                             arcname=os.path.basename(hydro_paths[fmt]))
                if fmt in bonjean_paths:
                    zf.write(bonjean_paths[fmt],
                             arcname=os.path.basename(bonjean_paths[fmt]))

        return send_file(
            zip_path,
            as_attachment=True,
            download_name=f"曲线图_{timestamp}.zip",
            mimetype="application/zip"
        )

    except Exception as e:
        return jsonify({
            "success": False,
            "msg": f"曲线导出失败：{str(e)}",
            "trace": traceback.format_exc()
        })


@app.route("/download_plot/<path:filename>")
def download_plot(filename):
    """
    【阶段4新增】下载单个曲线文件（PNG 或 SVG）。
    """
    safe_name = os.path.basename(filename)
    file_path = os.path.join(PLOT_DIR, safe_name)
    if not os.path.exists(file_path):
        return jsonify({"success": False, "msg": "文件不存在"}), 404
    return send_file(file_path, as_attachment=True,
                     download_name=safe_name)


# ══════════════════════════════════════════════════════════
# 阶段5 新增路由：GZ 曲线计算与绘制
# ══════════════════════════════════════════════════════════

@app.route("/calc_gz", methods=["POST"])
def calc_gz():
    """
    【阶段5新增】计算单工况 GZ 曲线（稳性横截曲线）。

    请求体（JSON）：
        {
            "offsets_data": { stations, waterlines, offsets },
            "draft":      5.0,    // 设计吃水（m）
            "KG":         4.5,    // 重心高度（m）
            "theta_min":  0.0,    // 最小横倾角（度），默认 0
            "theta_max":  90.0,   // 最大横倾角（度），默认 90
            "theta_step": 5.0,    // 步长（度），默认 5
            "rho":        1.025   // 水密度，默认海水
        }

    返回（JSON）：
        {
            "success": true,
            "gz": {
                "rows":          [...],   // 每个角度的 GZ 值
                "GM":            0.xxx,
                "GZ_max":        0.xxx,
                "theta_max_gz":  xx.x,
                "theta_vanish":  xx.x,
                "criteria":      {...},   // 稳性衡准校核
                ...
            }
        }
    """
    try:
        data = request.get_json()
        offsets_data = data.get("offsets_data")
        if not offsets_data:
            return jsonify({"success": False, "msg": "缺少型值表数据"})

        draft      = float(data.get("draft", 5.0))
        KG         = float(data.get("KG", 4.0))
        theta_min  = float(data.get("theta_min", 0.0))
        theta_max  = float(data.get("theta_max", 90.0))
        theta_step = float(data.get("theta_step", 5.0))
        rho        = float(data.get("rho", 1.025))

        gz_result = calc_gz_curve(
            offsets_data=offsets_data,
            draft=draft,
            KG=KG,
            theta_min=theta_min,
            theta_max=theta_max,
            theta_step=theta_step,
            rho=rho,
        )

        return jsonify({"success": True, "gz": gz_result})

    except Exception as e:
        return jsonify({
            "success": False,
            "msg": f"GZ 计算出错：{str(e)}",
            "trace": traceback.format_exc()
        })


@app.route("/calc_gz_family", methods=["POST"])
def calc_gz_family_route():
    """
    【阶段5新增】计算多排水量 GZ 曲线族。

    请求体（JSON）：
        {
            "offsets_data": { ... },
            "drafts":     [3.0, 4.0, 5.0, 6.0],  // 吃水列表
            "KG":         4.5,
            "theta_step": 5.0
        }
    """
    try:
        data = request.get_json()
        offsets_data = data.get("offsets_data")
        drafts       = [float(d) for d in data.get("drafts", [])]
        KG           = float(data.get("KG", 4.0))
        theta_step   = float(data.get("theta_step", 5.0))
        rho          = float(data.get("rho", 1.025))

        if not drafts:
            return jsonify({"success": False, "msg": "缺少吃水列表"})

        family = calc_gz_family(offsets_data, drafts, KG, theta_step, rho)
        return jsonify({"success": True, "family": family})

    except Exception as e:
        return jsonify({
            "success": False,
            "msg": f"GZ 曲线族计算出错：{str(e)}",
            "trace": traceback.format_exc()
        })


@app.route("/preview_gz", methods=["POST"])
def preview_gz():
    """
    【阶段5新增】生成 GZ 曲线预览图（base64 PNG）。

    请求体（JSON）：{ "gz": { ... } }
    """
    try:
        data = request.get_json()
        gz   = data.get("gz")
        if not gz:
            return jsonify({"success": False, "msg": "缺少 GZ 数据"})

        b64 = plot_gz_preview(gz)
        return jsonify({"success": True, "image": b64})

    except Exception as e:
        return jsonify({
            "success": False,
            "msg": f"预览生成失败：{str(e)}",
            "trace": traceback.format_exc()
        })


@app.route("/export_gz", methods=["POST"])
def export_gz():
    """
    【阶段5新增】导出 GZ 曲线图（PNG+SVG）和 GZ 数据表（Excel），打包为 ZIP。

    请求体（JSON）：
        {
            "gz":        { ... },   // 单工况 GZ 数据（必须）
            "family":    [ ... ],   // 多排水量 GZ 族（可选）
            "ship_name": "示例船"
        }
    """
    try:
        import zipfile
        data      = request.get_json()
        gz        = data.get("gz")
        family    = data.get("family")
        ship_name = data.get("ship_name", "")

        if not gz:
            return jsonify({"success": False, "msg": "缺少 GZ 数据"})

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        plot_dir  = os.path.join(PLOT_DIR, f"gz_{timestamp}")
        os.makedirs(plot_dir, exist_ok=True)

        # ── 绘制 GZ 曲线图 ────────────────────────────────
        gz_paths = plot_gz_curve(gz, plot_dir, ship_name=ship_name)

        # ── 绘制 GZ 曲线族（如果有）─────────────────────
        family_paths = {}
        if family:
            family_paths = plot_gz_family(family, plot_dir, ship_name=ship_name)

        # ── 导出 GZ 数据表（Excel）───────────────────────
        excel_path = _export_gz_excel(gz, plot_dir)

        # ── 打包 ZIP ──────────────────────────────────────
        zip_path = os.path.join(OUTPUT_DIR, f"GZ曲线_{timestamp}.zip")
        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zf:
            for p in [gz_paths.get('png'), gz_paths.get('svg'),
                      family_paths.get('png'), family_paths.get('svg'),
                      excel_path]:
                if p and os.path.exists(p):
                    zf.write(p, arcname=os.path.basename(p))

        return send_file(
            zip_path,
            as_attachment=True,
            download_name=f"GZ曲线_{timestamp}.zip",
            mimetype="application/zip"
        )

    except Exception as e:
        return jsonify({
            "success": False,
            "msg": f"GZ 导出失败：{str(e)}",
            "trace": traceback.format_exc()
        })


def _export_gz_excel(gz_result: dict, output_dir: str) -> str:
    """导出 GZ 曲线数据表为 Excel"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "GZ曲线数据"

    # 标题
    ws.merge_cells("A1:H1")
    c = ws.cell(1, 1, "稳性横截曲线（GZ 曲线）计算表")
    c.font = Font(name="宋体", size=12, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="1a4a8a")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # 基本参数
    params = [
        ("吃水 d", gz_result["draft"], "m"),
        ("KG",     gz_result["KG"],    "m"),
        ("KB",     gz_result["KB"],    "m"),
        ("BM",     gz_result["BM"],    "m"),
        ("KM",     gz_result["KB"]+gz_result["BM"], "m"),
        ("GM",     gz_result["GM"],    "m"),
        ("排水量", gz_result["delta"], "t"),
        ("GZ_max", gz_result["GZ_max"], "m"),
        ("θ_max",  gz_result["theta_max_gz"], "°"),
    ]
    for i, (name, val, unit) in enumerate(params):
        ws.cell(2, i*1+1, f"{name}={val:.3f}{unit}")

    # 数据表头
    headers = ["横倾角θ(°)", "GZ(m)", "GZ近似(m)", "稳性横截臂ls(m)",
               "浮心横坐标yB(m)", "浮心垂坐标zB(m)"]
    for j, h in enumerate(headers):
        c = ws.cell(4, j+1, h)
        c.font = Font(name="宋体", size=9, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="2e7dd1")
        c.alignment = Alignment(horizontal="center")

    # 数据行
    for i, row in enumerate(gz_result["rows"]):
        r = i + 5
        ws.cell(r, 1, row["theta"])
        ws.cell(r, 2, row["GZ"])
        ws.cell(r, 3, row["GZ_approx"])
        ws.cell(r, 4, row["ls"])
        ws.cell(r, 5, row["y_B"])
        ws.cell(r, 6, row["z_B"])

    # 稳性衡准
    crit_row = len(gz_result["rows"]) + 7
    ws.cell(crit_row, 1, "稳性衡准校核（《国内航行海船法定检验技术规则》）")
    ws.cell(crit_row, 1).font = Font(bold=True, color="1a4a8a")

    criteria = gz_result["criteria"]
    for k, (key, crit) in enumerate(criteria.items()):
        if key == "overall" or not isinstance(crit, dict):
            continue
        r = crit_row + 1 + k
        ok = crit.get("pass", False)
        ws.cell(r, 1, crit.get("name", key))
        ws.cell(r, 2, f"{crit.get('value', '')}")
        ws.cell(r, 3, crit.get("unit", ""))
        ws.cell(r, 4, f">={crit.get('limit', '')}")
        ws.cell(r, 5, "通过" if ok else "不通过")
        ws.cell(r, 5).font = Font(color="059669" if ok else "d62728", bold=True)

    for col in range(1, 7):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 16

    path = os.path.join(output_dir, "GZ曲线数据.xlsx")
    wb.save(path)
    return path


# ══════════════════════════════════════════════════════════
# 阶段6 路由（装载工况稳性校核与规范判定）
# ══════════════════════════════════════════════════════════

@app.route("/list_loading_conditions", methods=["GET"])
def list_loading_conditions():
    """
    【阶段6新增】列出所有装载工况
    
    返回：
        {
            "success": true,
            "conditions": ["满载出港", "满载到港", "压载出港", "压载到港", ...]
        }
    """
    try:
        manager = LoadingConditionManager()
        conditions = manager.list_conditions()
        
        return jsonify({
            "success": True,
            "conditions": conditions
        })
    
    except Exception as e:
        return jsonify({
            "success": False,
            "msg": f"获取工况列表失败：{str(e)}"
        })


@app.route("/get_loading_condition", methods=["POST"])
def get_loading_condition():
    """
    【阶段6新增】获取指定装载工况的详细信息
    
    请求体：
        {
            "condition_name": "满载出港"
        }
    
    返回：
        {
            "success": true,
            "condition": {
                "name": "满载出港",
                "total_weight": 5200,
                "xg": 42.5,
                "yg": 0,
                "zg": 5.1,
                "compartments": {...}
            }
        }
    """
    try:
        data = request.get_json()
        condition_name = data.get("condition_name")
        
        if not condition_name:
            return jsonify({"success": False, "msg": "缺少工况名称"})
        
        manager = LoadingConditionManager()
        lc = manager.get_condition(condition_name)
        
        if lc is None:
            return jsonify({"success": False, "msg": f"工况 {condition_name} 不存在"})
        
        return jsonify({
            "success": True,
            "condition": lc.to_dict()
        })
    
    except Exception as e:
        return jsonify({
            "success": False,
            "msg": f"获取工况失败：{str(e)}",
            "trace": traceback.format_exc()
        })


@app.route("/analyze_loading_condition", methods=["POST"])
def analyze_loading_condition():
    """
    【阶段6新增】分析单个装载工况的稳性
    
    请求体：
        {
            "offsets_data": {...},
            "hydrostatics_table": {...},
            "condition_name": "满载出港",
            "draft": 5.8,
            "KG": 5.1
        }
    
    返回：
        {
            "success": true,
            "analysis": {
                "condition_name": "满载出港",
                "loading": {...},
                "floating_state": {...},
                "indicators": {...}
            },
            "judgment": {
                "overall_pass": true,
                "judgments": {...},
                "failed_items": []
            }
        }
    """
    try:
        data = request.get_json()
        offsets_data = data.get("offsets_data")
        hydrostatics_table = data.get("hydrostatics_table")
        condition_name = data.get("condition_name")
        
        if not all([offsets_data, hydrostatics_table, condition_name]):
            return jsonify({"success": False, "msg": "缺少必要参数"})
        
        # 获取工况
        manager = LoadingConditionManager()
        lc = manager.get_condition(condition_name)
        
        if lc is None:
            return jsonify({"success": False, "msg": f"工况 {condition_name} 不存在"})
        
        # 计算稳性
        stability_calc = LoadingConditionStability(
            lc, offsets_data, hydrostatics_table, calc_gz_curve
        )
        analysis_result = stability_calc.calculate_stability()
        
        # 规范判定
        judgment = StabilityJudgment(analysis_result)
        judgment_result = judgment.judge_all()
        
        return jsonify({
            "success": True,
            "analysis": analysis_result,
            "judgment": judgment_result
        })
    
    except Exception as e:
        return jsonify({
            "success": False,
            "msg": f"工况分析失败：{str(e)}",
            "trace": traceback.format_exc()
        })


@app.route("/analyze_all_loading_conditions", methods=["POST"])
def analyze_all_loading_conditions():
    """
    【阶段6新增】分析所有 4 种必做工况
    
    请求体：
        {
            "offsets_data": {...},
            "hydrostatics_table": {...}
        }
    
    返回：
        {
            "success": true,
            "results": {
                "满载出港": {...},
                "满载到港": {...},
                "压载出港": {...},
                "压载到港": {...}
            },
            "summary": {
                "total": 4,
                "passed": 3,
                "failed": 1
            }
        }
    """
    try:
        data = request.get_json()
        offsets_data = data.get("offsets_data")
        hydrostatics_table = data.get("hydrostatics_table")
        
        if not all([offsets_data, hydrostatics_table]):
            return jsonify({"success": False, "msg": "缺少必要参数"})
        
        # 创建分析器
        analysis = LoadingStabilityAnalysis(
            offsets_data, hydrostatics_table, calc_gz_curve
        )
        
        # 分析所有工况
        analysis.analyze_all_standard_conditions()
        
        # 获取判定总结
        summary = analysis.get_judgment_summary()
        
        # 组织结果
        results = {}
        for condition_name in ['满载出港', '满载到港', '压载出港', '压载到港']:
            if condition_name in analysis.analysis_results:
                results[condition_name] = {
                    'analysis': analysis.analysis_results[condition_name],
                    'judgment': analysis.judgment_results[condition_name]
                }
        
        return jsonify({
            "success": True,
            "results": results,
            "summary": {
                "total": summary['total_conditions'],
                "passed": summary['passed_conditions'],
                "failed": summary['failed_conditions']
            }
        })
    
    except Exception as e:
        return jsonify({
            "success": False,
            "msg": f"全工况分析失败：{str(e)}",
            "trace": traceback.format_exc()
        })


@app.route("/export_stability_report", methods=["POST"])
def export_stability_report():
    """
    【阶段6新增】导出稳性校核报告 Excel
    
    请求体：
        {
            "condition_name": "满载出港",
            "analysis": {...},
            "judgment": {...}
        }
    
    返回：
        {
            "success": true,
            "file": "稳性校核报告_满载出港.xlsx",
            "download_url": "/download_plot/稳性校核报告_满载出港.xlsx"
        }
    """
    try:
        data = request.get_json()
        condition_name = data.get("condition_name")
        analysis = data.get("analysis")
        judgment = data.get("judgment")
        
        if not all([condition_name, analysis, judgment]):
            return jsonify({"success": False, "msg": "缺少必要参数"})
        
        # 创建导出器
        exporter = StabilityReportExporter()
        
        # 生成文件名
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"稳性校核报告_{condition_name}_{timestamp}.xlsx"
        output_path = os.path.join(OUTPUT_DIR, filename)
        
        # 导出报告
        exporter.create_single_condition_report(condition_name, analysis, judgment, output_path)
        
        return jsonify({
            "success": True,
            "file": filename,
            "download_url": f"/download_plot/{filename}"
        })
    
    except Exception as e:
        return jsonify({
            "success": False,
            "msg": f"导出失败：{str(e)}",
            "trace": traceback.format_exc()
        })


@app.route("/export_all_conditions_report", methods=["POST"])
def export_all_conditions_report():
    """
    【阶段6新增】导出全工况稳性校核报告 Excel
    
    请求体：
        {
            "results": {
                "满载出港": {analysis, judgment},
                "满载到港": {analysis, judgment},
                ...
            }
        }
    
    返回：
        {
            "success": true,
            "file": "全工况稳性校核报告.xlsx",
            "download_url": "/download_plot/全工况稳性校核报告.xlsx"
        }
    """
    try:
        data = request.get_json()
        results = data.get("results")
        
        if not results:
            return jsonify({"success": False, "msg": "缺少结果数据"})
        
        # 创建导出器
        exporter = StabilityReportExporter()
        
        # 生成文件名
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"全工况稳性校核报告_{timestamp}.xlsx"
        output_path = os.path.join(OUTPUT_DIR, filename)
        
        # 导出报告
        exporter.create_all_conditions_report(results, output_path)
        
        return jsonify({
            "success": True,
            "file": filename,
            "download_url": f"/download_plot/{filename}"
        })
    
    except Exception as e:
        return jsonify({
            "success": False,
            "msg": f"导出失败：{str(e)}",
            "trace": traceback.format_exc()
        })


# ══════════════════════════════════════════════════════════
# 阶段7：课程设计 Word 说明书自动生成
# ══════════════════════════════════════════════════════════

@app.route("/generate_word_report", methods=["POST"])
def generate_word_report():
    """
    【阶段7新增】生成课程设计 Word 说明书
    
    请求体：
        {
            "user_info": {
                "class_name": "船舶2020班",
                "student_name": "张三",
                "student_id": "202012345",
                "instructor": "李教授"
            },
            "ship_data": {
                "ship_name": "3000DWT沿海散货船",
                "LOA": 98.0,
                "LBP": 92.0,
                ...
            },
            "hydrostatics_result": {...},
            "bonjean_result": {...},
            "stability_results": {...},
            "image_paths": {
                "hydrostatics": "静水力曲线.png",
                "bonjean": "邦戎曲线.png",
                "gz_curve": "GZ曲线.png"
            }
        }
    
    返回：
        {
            "success": true,
            "file": "课程设计说明书_张三_202012345.docx",
            "download_url": "/download_plot/课程设计说明书_张三_202012345.docx"
        }
    """
    try:
        data = request.get_json()
        
        user_info = data.get("user_info", {})
        ship_data = data.get("ship_data", {})
        hydrostatics_result = data.get("hydrostatics_result", {})
        bonjean_result = data.get("bonjean_result", {})
        stability_results = data.get("stability_results", {})
        image_filenames = data.get("image_paths", {})
        
        # 转换图片文件名为完整路径
        image_paths = {}
        for key, filename in image_filenames.items():
            if filename:
                full_path = os.path.join(OUTPUT_DIR, filename)
                if os.path.exists(full_path):
                    image_paths[key] = full_path
        
        # 创建生成器
        generator = CourseDesignReportGenerator()
        
        # 生成文件名
        student_name = user_info.get("student_name", "学生")
        student_id = user_info.get("student_id", "学号")
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"课程设计说明书_{student_name}_{student_id}_{timestamp}.docx"
        output_path = os.path.join(OUTPUT_DIR, filename)
        
        # 生成报告
        generator.generate_complete_report(
            user_info, ship_data, hydrostatics_result, bonjean_result,
            stability_results, image_paths, output_path
        )
        
        return jsonify({
            "success": True,
            "file": filename,
            "download_url": f"/download_plot/{filename}"
        })
    
    except Exception as e:
        return jsonify({
            "success": False,
            "msg": f"生成失败：{str(e)}",
            "trace": traceback.format_exc()
        })


# ── 启动 ─────────────────────────────────────────────────
if __name__ == "__main__":
    print("=" * 60)
    print("  船舶静力学课程设计系统  —  武汉理工大学船海学院")
    print("  阶段4：静水力曲线 + 邦戎曲线绘制（A3/300DPI）")
    print("  阶段5：GZ 曲线（稳性横截曲线）计算与绘图")
    print("  访问地址：http://127.0.0.1:5000")
    print("=" * 60)
    app.run(debug=True, port=5000, use_reloader=False)
